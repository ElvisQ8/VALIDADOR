# PROYECTO ELVIS Q.
import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import plotly.graph_objects as go
from io import BytesIO

import os
from openpyxl.styles import PatternFill, Font
import plotly.express as px
import csv
import io
# CERTIFICADO
def procesar_archivo(archivo_cargado, plantilla):
   
    plantilla_wb = openpyxl.load_workbook(plantilla)
    plantilla_ws = plantilla_wb["PECLD07792"]
    duplicado_ws = plantilla_wb["Duplicado"]
    standar_ws = plantilla_wb["STD (PECLSTDEN02)"]

    wb = openpyxl.load_workbook(archivo_cargado, data_only=True)
    if "BD_densidad_2020" not in wb.sheetnames:
        st.error("El archivo cargado no contiene la hoja 'BD_densidad_2020'")
        return None

    ws = wb["BD_densidad_2020"]
    datos = [row for row in ws.iter_rows(min_row=10, max_col=17, values_only=True) if any(row)]

    start_row = 28
    for i, row in enumerate(datos, start=start_row):
        for j, value in enumerate(row, start=1):
            plantilla_ws.cell(row=i, column=j, value=value)

    from openpyxl.styles import PatternFill
    fill = PatternFill(start_color="E26B0A", end_color="E26B0A", fill_type="solid")
    for row in plantilla_ws.iter_rows(min_row=28, max_col=17):
        if any(cell.value in ["DSTD", "DEND"] for cell in row):
            for cell in row:
                cell.fill = fill

    dest_row = 11
    for row in plantilla_ws.iter_rows(min_row=27, min_col=15, max_col=15, values_only=False):
        if row[0].value == "DEND":
            fila_actual = row[0].row
            duplicado_ws.cell(row=dest_row, column=1, value=plantilla_ws.cell(row=fila_actual, column=4).value)
            duplicado_ws.cell(row=dest_row, column=3, value=plantilla_ws.cell(row=fila_actual, column=4).value)
            duplicado_ws.cell(row=dest_row, column=4, value=plantilla_ws.cell(row=fila_actual, column=13).value)
            duplicado_ws.cell(row=dest_row, column=2, value=plantilla_ws.cell(row=fila_actual - 1, column=13).value)
            dest_row += 1

    valor_d1 = plantilla_ws.cell(row=28, column=17).value
    valor_d2 = plantilla_ws.cell(row=28, column=13).value

    
    final_data_row = start_row + len(datos)
    plantilla_ws.delete_rows(final_data_row + 0, 80)
    standar_ws.cell(row=11, column=2, value=valor_d1)
    standar_ws.cell(row=11, column=4, value=valor_d2)

    nuevo_nombre = plantilla_ws.cell(row=28, column=1).value
    if nuevo_nombre:
        plantilla_ws.title = str(nuevo_nombre)

    output = BytesIO()
    plantilla_wb.save(output)
    output.seek(0)
    return output

# INTERFAZ 
st.title("OPTIMIZACION DE PROCESOS - DENSIDADES")
pagina = st.sidebar.radio("Selecciona un proceso", ["Generar certificado", "Exportador", "YAMILA"])

if pagina == "Generar certificado":
    st.subheader("Generación de certificado")

    opciones_plantilla = {
        "ARTURO": "PLANTILLA.xlsx",
        "MILAGROS ": "PLANTILLA1.xlsx",
        "YONATAN": "PLANTILLA2.xlsx"
    }
    plantilla_seleccionada = st.selectbox("Seleccione el responsable:", list(opciones_plantilla.keys()))
    plantilla_path = opciones_plantilla[plantilla_seleccionada]

    archivo_cargado = st.file_uploader("Carga archivo de datos en Excel", type=["xlsx"])

    if archivo_cargado:
        # certificado
        output = procesar_archivo(archivo_cargado, plantilla_path)
        if output:
            st.download_button("Descargar archivo procesado", data=output, file_name="Certificado.xlsx")

        # ANÁLISIS DE DENSIDADES CL
        st.divider()
        st.subheader("Análisis de Densidades")

        df = pd.read_excel(archivo_cargado, sheet_name=0, header=None)
        df = df.drop(index=np.arange(8)).reset_index(drop=True)
        df.columns = df.iloc[0]
        df = df.drop(index=0).reset_index(drop=True)

        
        df['TIPO DE CONTROL QA/QC'] = df['TIPO DE CONTROL QA/QC'].fillna('ORD')
        df['MUESTRA'] = df['MUESTRA'].fillna('ESTANDAR')

        
        metodo = st.multiselect("Filtrar por MÉTODO DE ANÁLISIS", sorted(df['MÉTODO DE ANÁLISIS'].dropna().unique()))
        tipo_control = st.multiselect("Filtrar por TIPO DE CONTROL QA/QC", sorted(df['TIPO DE CONTROL QA/QC'].dropna().unique()))
        comentario = st.multiselect("Filtrar por DOMINIO", sorted(df['COMENTARIO'].dropna().unique()))

        filtrado = df.copy()
        if metodo:
            filtrado = filtrado[filtrado['MÉTODO DE ANÁLISIS'].isin(metodo)]
        if tipo_control:
            filtrado = filtrado[filtrado['TIPO DE CONTROL QA/QC'].isin(tipo_control)]
        if comentario:
            filtrado = filtrado[filtrado['COMENTARIO'].isin(comentario)]

        # LIMITES DE DATA ENTREGADA POR LOS TECNICOS ( ESTO QUEDA PENDIENTE REAFIRMAR CON MAS ENSAYOS)
        rangos_lito = {
            'D': (2.67, 2.8), 'D1': (2.71, 2.95), 'VD': (2.51, 3.26), 'VM': (2.55, 3.86),
            'SSM': (2.8, 4.2), 'SPB': (3.32, 4.94), 'SPP': (3.51, 4.9),
            'PECLSTDEN02': (2.749, 2.779), 'SSL': (2.8, 4.2), 'SOB': (3.32, 4.94),
            'SOP': (3.51, 4.9), 'VL': (2.51, 3.26)
        }

        estados, comentarios = [], []
        for idx, row in filtrado.iterrows():
            densidad = row['DENSIDAD']
            litologia = row['COMENTARIO']
            if pd.isna(densidad):
                estados.append('Sin Densidad')
                comentarios.append('')
                continue
            if pd.isna(litologia):
                if 2.749 <= densidad <= 2.779:
                    estados.append('Correcto')
                    comentarios.append('Estándar dentro del rango')
                else:
                    estados.append('Fuera de Rango')
                    comentarios.append('Estándar fuera de rango')
            elif litologia in rangos_lito:
                min_val, max_val = rangos_lito[litologia]
                estados.append('Correcto' if min_val <= densidad <= max_val else 'Fuera de Rango')
                comentarios.append('')
            else:
                estados.append('Litología desconocida')
                comentarios.append('')

        filtrado['Estado'] = estados
        filtrado['Comentario Validación'] = comentarios

        
        for idx in range(1, len(filtrado)):
            if filtrado.iloc[idx]['TIPO DE CONTROL QA/QC'] == 'DEND':
                actual = filtrado.iloc[idx]['DENSIDAD']
                anterior = filtrado.iloc[idx - 1]['DENSIDAD']
                if pd.notna(actual) and pd.notna(anterior):
                    var_pct = abs(actual - anterior) / anterior
                    if var_pct > 0.10:
                        filtrado.at[idx, 'Estado'] = 'Error Duplicado'
                        filtrado.at[idx, 'Comentario Validación'] = 'Duplicado fuera del 10%'
                    else:
                        filtrado.at[idx, 'Comentario Validación'] = 'Duplicado dentro del 10%'

        
        st.dataframe(filtrado)

        
        fig = go.Figure()
        for lit, (min_v, max_v) in rangos_lito.items():
            fig.add_shape(type="line", x0=0, x1=len(filtrado), y0=min_v, y1=min_v, line=dict(color="gray", dash="dash"))
            fig.add_shape(type="line", x0=0, x1=len(filtrado), y0=max_v, y1=max_v, line=dict(color="gray", dash="dash"))

        fig.add_trace(go.Scatter(
            x=filtrado['MUESTRA'],
            y=filtrado['DENSIDAD'],
            mode='markers',
            marker=dict(
                color=np.where(filtrado['Estado'].isin(['Fuera de Rango', 'Error Duplicado']), 'red', 'blue'),
                size=8
            ),
            text=filtrado['COMENTARIO'],
            hovertemplate='<b>Muestra:</b> %{x}<br><b>Densidad:</b> %{y}<br><b>Litología:</b> %{text}<extra></extra>'
        ))

        fig.update_layout(title='Validación de Densidades', xaxis_title='MUESTRA', yaxis_title='Densidad')
        st.plotly_chart(fig)

elif pagina == "Exportador":
    st.subheader("Exportador de datos para FUSION")

    # EXPORTAR
    def load_data(file_path):
        return pd.read_excel(file_path, sheet_name=0, header=None, skiprows=27, usecols="A:R", nrows=101)

    def clean_data(df):
        return df[df != 'hola']

    def copy_data_to_template(df, sheet_name, selected_name, template_file):
        template = pd.ExcelFile(template_file)
        with BytesIO() as output:
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                if sheet_name in template.sheet_names:
                    temp_df = template.parse(sheet_name)
                    temp_df.to_excel(writer, sheet_name=sheet_name, index=False)

                if sheet_name == "O":
                    df_filtered = df[~df[14].astype(str).str.contains('DSTD|DEND', na=False)]
                    df_filtered[13] = selected_name
                    columns_mapping_o = [(0, 'A'), (1, 'B'), (2, 'C'), (3, 'D'), (4, 'E'),
                                         (5, 'F'), (6, 'G'), (7, 'H'), (8, 'I'), (9, 'J'),
                                         (10, 'K'), (11, 'L'), (13, 'P'), (16, 'O')]
                    for df_col, template_col in columns_mapping_o:
                        writer.sheets[sheet_name].write_column(f'{template_col}2', df_filtered[df_col].fillna('').astype(str).values)

                elif sheet_name == "DP":
                    df_filtered = df[df[14].astype(str).str.contains('DEND', na=False)]
                    df_filtered[13] = selected_name
                    columns_mapping_dp = [(0, 'A'), (1, 'B'), (2, 'C'), (3, 'D'), (4, 'E'),
                                          (6, 'F'), (7, 'G'), (8, 'H'), (9, 'I'), (10, 'J'),
                                          (14, 'K'), (13, 'N'), (16, 'M'), (17, 'O')]
                    for df_col, template_col in columns_mapping_dp:
                        writer.sheets[sheet_name].write_column(f'{template_col}2', df_filtered[df_col].fillna('').astype(str).values)

                elif sheet_name == "STD":
                    df_filtered = df[df[14].astype(str).str.contains('DSTD', na=False)]
                    df_filtered[13] = selected_name
                    columns_mapping_std = [(0, 'A'), (4, 'B'), (6, 'C'), (7, 'D'), (8, 'E'),
                                           (9, 'F'), (10, 'G'), (11, 'H'), (13, 'K'), (16, 'J'), (17, 'L')]
                    peclstd_value = "PECLSTDEN02"
                    writer.sheets[sheet_name].write_column('H2', [peclstd_value] * len(df_filtered))
                    for df_col, template_col in columns_mapping_std:
                        writer.sheets[sheet_name].write_column(f'{template_col}2', df_filtered[df_col].fillna('').astype(str).values)

            output.seek(0)
            df_csv = pd.read_excel(output, sheet_name=sheet_name)
            csv_output = BytesIO()
            df_csv.to_csv(csv_output, index=False, sep=';', encoding='utf-8')
            csv_output.seek(0)
            return csv_output.getvalue()

    names = ["AJGU", "MIAP", "RYSA"]
    selected_name = st.selectbox("Selecciona un usuario", names)
    uploaded_file = st.file_uploader("Cargar el certificado .xlsx", type=["xlsx"])
    template_file = "plantilla_export.xlsx"

    if uploaded_file is not None:
        df = load_data(uploaded_file)

        if st.button('Exportar Hoja O'):
            df_cleaned = clean_data(df)
            file_o = copy_data_to_template(df_cleaned, "O", selected_name, template_file)
            st.download_button("Descargar Hoja O como CSV", data=file_o, file_name="plantilla_O.csv")

        if st.button('Exportar Hoja DP'):
            df_cleaned = clean_data(df)
            file_dp = copy_data_to_template(df_cleaned, "DP", selected_name, template_file)
            st.download_button("Descargar Hoja DP como CSV", data=file_dp, file_name="plantilla_DP.csv")

        if st.button('Exportar Hoja STD'):
            df_cleaned = clean_data(df)
            file_std = copy_data_to_template(df_cleaned, "STD", selected_name, template_file)
            st.download_button("Descargar Hoja STD como CSV", data=file_std, file_name="plantilla_STD.csv")
elif pagina == "YAMILA":
    st.subheader("VALIDADOR DE LOGUEO")
    def leer_csv(archivo):
        try:
            if archivo is None:
                st.error("Error: No se ha subido ningún archivo.")
                return None

            # Intentar primero con utf-8
            try:
                archivo.seek(0)
                df = pd.read_csv(archivo, encoding="utf-8", on_bad_lines="skip")
            except UnicodeDecodeError:
                archivo.seek(0)
                df = pd.read_csv(archivo, encoding="utf-16", on_bad_lines="skip")

            if df.empty:
                st.error(f"Error: El archivo {archivo.name} está vacío.")
                return None

            return df
        except Exception as e:
            st.error(f"Error al leer el archivo {archivo.name}: {e}")
            return None

    def validar_geo(df, hole_number):
        if df is None:
            st.error("⚠️ Error: No se pudo cargar el archivo Geology correctamente.")
            return None  # Detener la función si el DataFrame está vacío

        df.columns = df.columns.str.strip().str.lower()  # Normalizar nombres de columnas
        
        # Filtrar por HOLE_NUMBER
        df_filtrado = df[df["hole_number"] == hole_number]

        if df_filtrado.empty:
            st.warning(f"No se encontraron datos para HOLE_NUMBER: {hole_number}")
            return None

        condiciones = {
            31: ["VD"], 3: ["D", "D1"], 37: ["VAND"], 2: ["VL"], 28: ["VM"], 
            6: ["SPP"], 7: ["SOP"], 9: ["SPB"], 10: ["SOB"], 25: ["SSL"], 
            5: ["SSM"], 34: ["BXMM"], 30: ["I"], 14: ["P"], 8: ["BXC"], 
            32: ["VRD"], 33: ["VRD"], 12: ["CO"], 13: ["Q"], 17: ["LOST"], 15: ["F"]
        }

        df_filtrado['validación_geo'] = df_filtrado.apply(
            lambda row: 'correcto' if row['clito'] in condiciones and row['unit'] in condiciones[row['clito']] 
            else 'incorrecto', axis=1
        )

        return df_filtrado

    # Función para validar Sample y Standards

    def validar_sample_standards(sample_df, standards_df, hole_number):
        try:
            sample_df.columns = sample_df.columns.str.strip().str.lower()
            standards_df.columns = standards_df.columns.str.strip().str.lower()

            # Filtrar por HOLE_NUMBER
            sample_filtered = sample_df[sample_df['hole_number'] == hole_number].copy()
            standards_filtered = standards_df[standards_df['hole_number'] == hole_number].copy()

            # DataFrame para Sample
            df_sample = sample_filtered[['hole_number', 'sample_number', 'depth_from', 'depth_to', 'assay_sample_type_code','parent_sample_number']].copy()
            df_sample['tipo_muestra'] = df_sample['assay_sample_type_code']
            df_sample['depth_range'] = df_sample['depth_to'] - df_sample['depth_from']
            df_sample['tramo_valido'] = df_sample['depth_range'].apply(lambda x: '✅ Correcto' if 0.5 <= x <= 1.5 else '⚠️ Observado')
            df_sample = df_sample.drop(columns=['assay_sample_type_code'])

            # DataFrame para Standards (sin validación de tramo)
            df_standards = standards_filtered[['hole_number', 'sample_number', 'assay_standard_code']].copy()
            df_standards['tipo_muestra'] = df_standards['assay_standard_code']
            df_standards['depth_from'] = None
            df_standards['depth_to'] = None
            df_standards['depth_range'] = None
            df_standards['tramo_valido'] = None
            df_standards = df_standards.drop(columns=['assay_standard_code'])

            # Unir ambos DataFrames
            resultado = pd.concat([df_sample, df_standards], ignore_index=True)
            resultado = resultado.drop_duplicates(subset=['hole_number', 'sample_number', 'tipo_muestra'])
            resultado = resultado.sort_values(by='sample_number', ascending=True)

            # Reordenar columnas para visualización
            columnas_finales = [
                'hole_number', 'sample_number', 'tipo_muestra', 'parent_sample_number',
                'depth_from', 'depth_to', 'depth_range', 'tramo_valido'
            ]
            resultado = resultado[columnas_finales]

            return resultado
        except Exception as e:
            st.error(f"Error en validar_sample_standards: {e}")
            return None
        
    # Función para validar Alteration
    def validar_alteration(alteration_df, hole_number):
        try:
            if alteration_df is None:
                st.error("⚠️ Error: No se pudo cargar el archivo Alteration correctamente.")
                return None

            alteration_df.columns = alteration_df.columns.str.strip().str.lower()

            required_columns = ['hole_number', 'intensity_1', 'intensity_2', 'intensity_3', 
                                'distribution_1', 'distribution_2', 'distribution_3']
            missing_columns = [col for col in required_columns if col not in alteration_df.columns]

            if missing_columns:
                st.error(f"El archivo ALTERATION tiene columnas faltantes: {missing_columns}")
                return None

            # 🔹 Filtrar el DataFrame por `hole_number` antes de validar, asegurando que sea una copia independiente
            alteration_filtrado = alteration_df[alteration_df['hole_number'] == hole_number].copy()

            if alteration_filtrado.empty:
                st.warning(f"No se encontraron datos para HOLE_NUMBER: {hole_number}")
                return None

            def validar_filas(row):
                resultados = []
                for i in range(1, 4):
                    if row[f'intensity_{i}'] == 'FORT' and row[f'distribution_{i}'] != 'PERV':
                        resultados.append(f"Incorrecto en intensity_{i} y distribution_{i} (esperado PERV)")
                    if row[f'intensity_{i}'] == 'MODE' and pd.notnull(row[f'distribution_{i}']):
                        resultados.append(f"Incorrecto en intensity_{i} y distribution_{i} (esperado vacío)")
                    if row[f'intensity_{i}'] == 'FRCA' and row[f'distribution_{i}'] not in ['PUNT', 'VEIN']:
                        resultados.append(f"Incorrecto en intensity_{i} y distribution_{i} (esperado PUNT o VEIN)")

                return " | ".join(resultados) if resultados else "Correcto"

            # 🔹 Aplicar la validación **solo a las filas filtradas**
            alteration_filtrado.loc[:, 'validación'] = alteration_filtrado.apply(validar_filas, axis=1)

            # 🔹 Retornar SOLO el DataFrame filtrado
            return alteration_filtrado  
        except Exception as e:
            st.error(f"Error durante la validación en ALTERATION: {e}")
            return None

    # Mapeo entre Unit (Geology) y Rock_Type_Code (Major)
    correspondencias = {
        "D": "ANDS", "VAND": "ANDS", "D1": "DIOR", "VL": "DACT", "VM": "DACT", "VD": "DACT",
        "SPP": "MASS", "SOP": "MASS", "SPB": "MASS", "SOB": "MASS", "SSL": "MASS", "SSM": "SMSS",
        "BXMM": "FSTF", "I": "GRDR", "P": "PEGM", "BXC": "BRTC", "VRD": "RIDC", "CO": "SOLO",
        "Q": "VTQZ", "LOST": "XXXX", "F": "PNZO", "LOST": "YYYY"
    }


    # Función para validar intervalos
    def validar_intervalos(sample_df, validation_df, tipo, hole_number):
        try:
            sample_df.columns = sample_df.columns.str.strip().str.lower()
            validation_df.columns = validation_df.columns.str.strip().str.lower()

            sample_filtered = sample_df[sample_df['hole_number'] == hole_number]
            validation_filtered = validation_df[validation_df['hole_number'] == hole_number]

            if sample_filtered.empty:
                st.error(f"No se encontraron datos para HOLE_NUMBER {hole_number} en Sample.")
                return None
            if validation_filtered.empty:
                st.error(f"No se encontraron datos para HOLE_NUMBER {hole_number} en {tipo}.")
                return None

            sample_depth_from = sample_filtered['depth_from'].unique()
            sample_depth_to = sample_filtered['depth_to'].unique()

            resultados = []
            for _, row in validation_filtered.iterrows():
                depth_from_correcto = row['depth_from'] in sample_depth_from
                depth_to_correcto = row['depth_to'] in sample_depth_to

                validacion = "Correcto" if depth_from_correcto and depth_to_correcto else "Incorrecto"

                resultados.append({
                    'hole_number': row['hole_number'],
                    'depth_from': row['depth_from'],
                    'depth_to': row['depth_to'],
                    'archivo': tipo,
                    'validación': validacion
                })

            return pd.DataFrame(resultados)
        except Exception as e:
            st.error(f"Error durante la validación de intervalos en {tipo}: {e}")
            return None

    # Función para validar Major vs Geology
    def validar_major_geology(geology_df, major_df, hole_number):
        try:
            geology_df.columns = geology_df.columns.str.strip().str.lower()
            major_df.columns = major_df.columns.str.strip().str.lower()

            geology_filtered = geology_df[geology_df['hole_number'] == hole_number]
            major_filtered = major_df[major_df['hole_number'] == hole_number]

            if geology_filtered.empty:
                st.error(f"No se encontraron datos para HOLE_NUMBER {hole_number} en Geology.")
                return None
            if major_filtered.empty:
                st.error(f"No se encontraron datos para HOLE_NUMBER {hole_number} en Major.")
                return None

            resultados = []
            for _, major_row in major_filtered.iterrows():
                major_from, major_to, rock_type = major_row['depth_from'], major_row['depth_to'], major_row['rock_type_code']

                geology_segmentos = geology_filtered[
                    (geology_filtered['depth_from'] >= major_from) & (geology_filtered['depth_to'] <= major_to)
                ]

                if geology_segmentos.empty:
                    validacion = "Incorrecto (No contiene segmentos de Geology)"
                else:
                    unidades_geology = geology_segmentos['unit'].unique()
                    validacion = "Correcto" if all(correspondencias.get(unit, "") == rock_type for unit in unidades_geology) else "Incorrecto (Rock_Type no coincide con Units de Geology)"

                resultados.append({
                    'hole_number': hole_number,
                    'depth_from_major': major_from,
                    'depth_to_major': major_to,
                    'rock_type_major': rock_type,
                    'validación': validacion
                })

            return pd.DataFrame(resultados)
        except Exception as e:
            st.error(f"Error durante la validación entre Geology y Major: {e}")
            return None

    # Interfaz en Streamlit
    st.title("Validación de Datos Geológicos")

    hole_number = st.text_input("Ingrese el HOLE_NUMBER a buscar:", key="hole_number_input")

    # Carga de archivos en formato TXT y conversión a DataFrame
    geology_file = st.file_uploader("Cargar Geology (.csv)", type=[".csv"], key="geology_uploader")
    sample_file = st.file_uploader("Cargar Sample (.csv)", type=["csv"], key="sample_uploader")
    standards_file = st.file_uploader("Cargar Standards (.csv)", type=["csv"], key="standards_uploader")
    alteration_file = st.file_uploader("Cargar Alteration (.csv)", type=["csv"], key="alteration_uploader")
    mine_file = st.file_uploader("Cargar Mine (.csv)", type=["csv"], key="mine_uploader")
    major_file = st.file_uploader("Cargar Major (.csv)", type=["csv"], key="major_uploader")

    # Convertir archivos TXT a DataFrames
    geology_df = leer_csv(geology_file) if geology_file else None
    sample_df = leer_csv(sample_file) if sample_file else None
    standards_df = leer_csv(standards_file) if standards_file else None
    alteration_df = leer_csv(alteration_file) if alteration_file else None
    mine_df = leer_csv(mine_file) if mine_file else None
    major_df = leer_csv(major_file) if major_file else None


    # Función para descargar archivos
    def descargar_resultados(df, nombre_archivo):
        if df is not None and not df.empty:
            # Crear un buffer en memoria
            output = io.BytesIO()
            
            # Guardar el DataFrame en el buffer como archivo Excel
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False)
            
            # Mover el cursor al inicio del archivo
            output.seek(0)

            # Botón de descarga
            st.download_button(
                label=f"Descargar {nombre_archivo}",
                data=output,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    def exportar_a_excel(df, filename):
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Resultados', index=False)

                workbook = writer.book
                sheet = workbook['Resultados']

                # Formato del encabezado
                encabezado_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                encabezado_font = Font(bold=True)

                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.value = str(cell.value).upper()
                    cell.fill = encabezado_fill
                    cell.font = encabezado_font

                # Colores para tipo_muestra
                colores = {
                    "PECLSTD006": "F7F99F",
                    "PECLSTD007": "3785BF",
                    "RG": "F0DEF2",
                    "DP": "B5E6A2"
                }

                tipo_col = None
                for col in range(1, sheet.max_column + 1):
                    if sheet.cell(row=1, column=col).value == "TIPO_MUESTRA":
                        tipo_col = col
                        break

                if tipo_col:
                    for row in range(2, sheet.max_row + 1):
                        value = sheet.cell(row=row, column=tipo_col).value
                        if value in colores:
                            sheet.cell(row=row, column=tipo_col).fill = PatternFill(start_color=colores[value], end_color=colores[value], fill_type="solid")

                workbook.save(filename)

        except Exception as e:
            st.error(f"Error al exportar a Excel: {e}")
            
    # Botones de validación con tablas interactivas
    if st.button("Validar Geology", key="validate_geology") and geology_file:
        geology_df = leer_csv(geology_file)
        resultados_geo = validar_geo(geology_df, hole_number)
        st.dataframe(resultados_geo)  # Tabla interactiva
        descargar_resultados(resultados_geo, "resultados_geology.csv")

    if st.button("Validar Sample & Standards", key="validate_sample_standards") and sample_file and standards_file:
        sample_df = leer_csv(sample_file)
        standards_df = leer_csv(standards_file)
        resultados_sample_standards = validar_sample_standards(sample_df, standards_df, hole_number)

        if resultados_sample_standards is not None:
            st.dataframe(resultados_sample_standards)
            exportar_a_excel(resultados_sample_standards, "PECLD07.xlsx")

            with open("PECLD07.xlsx", "rb") as file:
                st.download_button(label="⬇️ Descargar Excel",
                                   data=file,
                                   file_name="PECLD07.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                
    if st.button("Validar Alteration", key="validate_alteration") and alteration_file:
        alteration_df = leer_csv(alteration_file)
        resultados_alteration = validar_alteration(alteration_df, hole_number)
        st.dataframe(resultados_alteration)
        descargar_resultados(resultados_alteration, "resultados_alteration.csv")

    if st.button("Validar Intervals", key="validate_intervals") and sample_file:
        sample_df = leer_csv(sample_file)
        
        archivos = {
            "Geology": geology_file,
            "Major": major_file,
            "Alteration": alteration_file,
            "Mine": mine_file
        }
        
        resultados_totales = []
        
        for tipo, archivo in archivos.items():
            if archivo:
                validation_df = leer_csv(archivo)
                resultados = validar_intervalos(sample_df, validation_df, tipo, hole_number)
                if resultados is not None:
                    resultados_totales.append(resultados)
        
        if resultados_totales:
            resultados_finales = pd.concat(resultados_totales, ignore_index=True)
            st.dataframe(resultados_finales)
            descargar_resultados(resultados_finales, "resultados_validacion.csv")

    if st.button("Validar Major", key="validate_major") and geology_file and major_file:
        geology_df = leer_csv(geology_file)
        major_df = leer_csv(major_file)

        if major_df is None or major_df.empty:
            st.error("Error: El archivo Major está vacío o no se pudo cargar correctamente.")
            st.stop()  # Detiene la ejecución sin errores

        resultados_major = validar_major_geology(geology_df, major_df, hole_number)
        if resultados_major is not None:
            st.dataframe(resultados_major)
            descargar_resultados(resultados_major, "resultados_major.csv")



    # Lista de estándares a analizar
    estandares_relevantes = ["PECLSTD006", "DP", "RG", "PECLSTD007", "PECLBLK002"]

    # Función para calcular el porcentaje de estándares respecto a las muestras OR
    def calcular_porcentaje_standards(sample_df, standards_df, hole_number):
        if sample_df is None or standards_df is None:
            st.warning("Error: No se han cargado ambos archivos correctamente.")
            return None, None

        # Filtrar por HOLE_NUMBER
        sample_filtrado = sample_df[sample_df["hole_number"] == hole_number]
        standards_filtrado = standards_df[standards_df["hole_number"] == hole_number]

        # Filtrar muestras OR (excluyendo DP y RG)
        muestras_or = sample_filtrado[~sample_filtrado["assay_sample_type_code"].isin(["DP", "RG"])]
        total_muestras_or = len(muestras_or)

        # Filtrar muestras DP y RG (que cuentan como estándares)
        muestras_dp_rg = sample_filtrado[sample_filtrado["assay_sample_type_code"].isin(["DP", "RG"])]
        total_dp_rg = len(muestras_dp_rg)

        total_standards = len(standards_filtrado)

        if total_muestras_or == 0 and total_standards == 0 and total_dp_rg == 0:
            st.warning(f"No se encontraron datos para HOLE_NUMBER: {hole_number}")
            return None, None

        # Calcular porcentaje con la ecuación correcta
        porcentaje_standards = (total_standards + total_dp_rg) / (total_standards + total_dp_rg + total_muestras_or) if (total_standards + total_dp_rg + total_muestras_or) > 0 else 0

        # Crear DataFrame con los resultados
        resumen_df = pd.DataFrame({
            "HOLE_NUMBER": [hole_number],
            "Total Muestras OR": [total_muestras_or],
            "Total Standards Relevantes": [total_standards],
            "Total DP/RG (como estándares)": [total_dp_rg],
            "Porcentaje Standards (%)": [porcentaje_standards * 100]
        })

        return resumen_df, porcentaje_standards


    # Botón para validar Sample & Standards y calcular el porcentaje
    if st.button("Ingreso de Sample & Standards", key="validate_sample_standards2") and hole_number:
        resultados_sample_standards = validar_sample_standards(sample_df, standards_df, hole_number)
        st.subheader("Resultados de validación:")
        st.dataframe(resultados_sample_standards)

        # 🔥 Nuevo análisis de porcentaje de estándares
        resumen_df, porcentaje = calcular_porcentaje_standards(sample_df, standards_df, hole_number)
        if resumen_df is not None:
            st.subheader("Resultados del análisis de estándares")
            st.dataframe(resumen_df)  # Tabla interactiva

            # Gráfico de barras con datos filtrados
            fig = px.bar(
                resumen_df.melt(value_vars=["Total Muestras OR", "Total Standards Relevantes", "Total DP/RG (como estándares)"]),
                x="variable", y="value", text="value",
                title=f"Comparación entre Muestras OR, Standards y DP/RG para HOLE_NUMBER {hole_number}",
                labels={"variable": "Tipo", "value": "Cantidad"},
                color="variable"
            )
            st.plotly_chart(fig)
